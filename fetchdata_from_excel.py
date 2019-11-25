import openpyxl
import json
import requests
import jsonpath

def test_add_multiple_students():
    api_url="http://thetestingworldapi.com/api/studentsDetails"
    f=open("C:/Users/Ranjitha/get_api.json", "r")
    wk=openpyxl.load_workbook("C:/Users/Ranjitha/data.xlsx")
    sh=wk['Sheet1']
    rows=sh.max_rowf
    request_json=json.loads(f.read())
    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i,column=1)
        cell_middle_name = sh.cell(row=i,column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        request_json['first_name']=cell_first_name.value
        request_json['middle_name']=cell_middle_name.value
        request_json['last_name']=cell_last_name.value
        request_json['date_of_birth']=cell_dob.value
        response=requests.post(api_url,request_json)
        print(response.status_code)
        assert response.status_code == 201